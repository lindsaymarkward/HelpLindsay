"""
Collate peer assessment forms for each group based on existing files (named systematically)
"""

import os
import openpyxl
from collections import namedtuple

__author__ = 'Lindsay Ward'

# Remove start (from LearnJCU)
WORKING_DIRECTORY = "/Users/sci-lmw1/Downloads/peer"
CAMPUSES = ["Ext", "Tsv", "Cns"]
FIRST_ROW = 8  # first row where student data appears -1 for 0-based index
MAX_GROUP_SIZE = 5
StudentRecord = namedtuple("StudentRecord", ["campus", "group", "firstname", "lastname"])

def main():
    students = makeStudentsDictionary()
    # cleanupDownloadedFiles(students)
    collateData(students)


def cleanupDownloadedFiles(students):
    os.chdir(WORKING_DIRECTORY)
    for filename in os.listdir('.'):
        # remove LearnJCU's text files
        if filename.endswith(".txt") or filename.startswith("."):
            os.remove(filename)
            continue

        # can only guarantee the part of the file name that LearnJCU prepends, e.g.
        #     Peer Assessment_jc232507_
        # cut the login from between the first two _

        login = filename[filename.find("_") + 1:filename.find("_attempt")]
        # print(login)
        # print(login, students[login][1])
        newName = "_".join([str(element) for element in students[login]]) + "-" + login + ".xlsx"
        # print(newName)

        os.rename(filename, newName)

        # Previous version - assumed students had not renamed files
        # get rid of LearnJCU's massive file name text and also the "attempt" text
        # file name createEvents:
        # Peer Assessment_jc232507_attempt_2015-10-30-10-21-30_Ext10AlexandraCameronPA.xlsx
        # remove from the last _ onwards
        # if "_" in filename:
        #     newName = filename[filename.rfind("_")+1:]
        #     os.rename(filename, newName)
        #     # print(newName)


def makeStudentsDictionary():
    students = {}  # dictionary with logins as keys
    groupsWB = openpyxl.load_workbook('/Users/sci-lmw1/GoogleDrive/Workspaces/Python/HelpLindsay/CP1406Groups.xlsx')
    campuses = groupsWB.get_sheet_names()

    for campus in campuses:
        groupsSheet = groupsWB.get_sheet_by_name(campus)

        for i in range(1, groupsSheet.max_row):
            groupCell, usernameCell, lastNameCell, firstNameCell = groupsSheet.rows[i]
            students[usernameCell.value] = StudentRecord(campus, groupCell.value, firstNameCell.value, lastNameCell.value)

    return students


def collateData(students):
    destinationFile = "/Users/sci-lmw1/GoogleDrive/Workspaces/Python/HelpLindsay/CollatedPeerAssessments.xlsx"
    destinationWorkbook = openpyxl.load_workbook(destinationFile)
    destinationSheet = destinationWorkbook.active

    os.chdir(WORKING_DIRECTORY)
    rowCounter = 2  # skip header row (0)
    # groupCounter = -1  # for adding blank rows between groups
    for filename in os.listdir('.'):
        # print(filename)
        if not filename.endswith(".xlsx"):
            continue
        # filename will be like: Townsville_1_Jack_Walsh-jc123456.xlsx
        login = filename[filename.rfind("-") + 1:filename.rfind(".")]
        student = students[login]
        print(student)
        sheet = openpyxl.load_workbook(filename).active
        for i in range(FIRST_ROW, FIRST_ROW + MAX_GROUP_SIZE + 1):
            # exit loop if we've run out of data (group is smaller than MAX_GROUP_SIZE)
            if not sheet.rows[i][0].value:
                break
            # print(sheet.rows[i][0].value)
            destinationSheet.cell(row=rowCounter, column=1).value = student.campus
            destinationSheet.cell(row=rowCounter, column=2).value = student.group
            destinationSheet.cell(row=rowCounter, column=3).value = student.firstname + " " + student.lastname

            # copy data for each group member
            for col in range(1, 11):
                destinationSheet.cell(row=rowCounter, column=col + 3).value = sheet.rows[i][col - 1].value

            rowCounter += 1
            # print(rowCounter, get_column_letter(rowCounter))

    destinationWorkbook.save(destinationFile)

main()
