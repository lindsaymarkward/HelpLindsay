"""Get CPXXXX assessment items from StudyFinder websites."""

import requests
import openpyxl
from collections import OrderedDict

YEAR = 2020
START_STRING = "<h3>Subject Assessment</h3>"
OUTPUT_HTML_FILENAME = "output/assessments.html"
OUTPUT_EXCEL_FILENAME = "output/2021-IT-Assessment-Mapping.xlsx"
HTML_TOP = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>IT@JCU Assessments in CSDB/Studyfinder</title>
    <style type="text/css">
        table, td {
            border: thin black solid;
            padding: 0.5em;
            border-collapse: collapse;
            vertical-align: top;
        }
    </style>
</head>
<body>
<h1>IT@JCU Assessments in CSDB/Studyfinder</h1>
"""
HTML_BOTTOM = """
</body>
</html>
"""


def main():
    all_subject_details = [OrderedDict(), OrderedDict()]  # assuming only 2 years
    # subjects = ['CP1401', 'CP1402']
    subjects = get_subjects()
    print(subjects)

    file_out = open(OUTPUT_HTML_FILENAME, 'w')
    print(HTML_TOP, file=file_out)
    subject_to_items = {}

    print(f"<table>", file=file_out)
    for subject in subjects:
        # Current year
        print(f"<tr>", file=file_out)
        for i in range(2):
            year_to_get = YEAR + i
            print(f"<td><h2>{subject} - {year_to_get}</h2>", file=file_out)
            url = f"https://secure.jcu.edu.au/app/studyfinder/index.cfm?subject={subject}&year={year_to_get}&transform=subjectwebview.xslt"
            response = requests.get(url)
            text = response.text
            assessment_block = get_assessment_block(text)
            print(assessment_block, file=file_out)
            items = extract_items(assessment_block)
            subject_to_items[subject] = items
            all_subject_details[i][subject] = items
            print(f"</td>", file=file_out)
        print(f"</tr>", file=file_out)
    print(f"</table>", file=file_out)
    print(HTML_BOTTOM, file=file_out)
    file_out.close()
    write_spreadsheet(all_subject_details)


def get_subjects():
    file_in = open("data/all_subjects.txt")
    subjects = [line.strip() for line in file_in]
    file_in.close()
    return subjects


def get_assessment_block(text):
    index_heading = text.find(START_STRING)
    index_end = text.find("</ul>", index_heading)
    section = text[index_heading + len(START_STRING):index_end + 6]
    section = section.replace('. ', '')
    return section.strip()


def extract_items(block):
    """Extract assessment items from HTML block as list of tuples."""
    items = []
    parts = block.split('\n')
    raw_items = [part.strip().strip('<li>').strip('</li>') for part in parts if part.strip().startswith('<li>')]
    for raw_item in raw_items:
        parts = raw_item.split(' - ')
        assessment = parts[0].replace('&gt;', '>')
        weight = parts[1].strip('(').strip('%)')
        items.append((assessment, weight))
    return items


def write_spreadsheet(all_subject_details):
    # all_subject_details contains 2 dictionaries
    workbook = openpyxl.load_workbook(filename=OUTPUT_EXCEL_FILENAME)
    sheet = workbook['Assessment-Mapping']
    row = 12  # first row for assessment items
    column = 2
    for i in range(2):  # for both years/dictionaries
        for subject, items in all_subject_details[i].items():
            # sheet.cell(row=row, column=column, value=subject)
            item_row = row
            for item_number, item in enumerate(items):
                name, weight = item
                try:
                    weight = int(weight)
                except ValueError:
                    pass
                sheet.cell(row=item_row + item_number, column=column, value=name)
                sheet.cell(row=item_row + item_number, column=column + 3, value=weight)
            column += 4  # distance to next subject (4 pieces of data per assessment)
        row += 7  # move down to write next year's items
        column = 2
    workbook.save(filename=f"output/temp.xlsx")


main()
