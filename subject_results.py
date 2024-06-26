"""
IT@JCU Subject Spreadsheet Solution
NO guarantees are given with this program :)
Lindsay Ward

Create subject results comprehensive spreadsheets, and optionally grade CSVs, from:
- Class list from Student Management System (SMS)
- LearnJCU Ultra grade centre download/export
- Blank subject results spreadsheet (template)

How To Use:
* Update user-customisable constants for desired directory and LearnJCU Grade Centre filename
  Creating the CSVs from the final grades (using xlwings) is by far the slowest
  set WILL_CREATE_CSV to False to skip this step (and do it yourself in Excel)
* Download LearnJCU Grade Centre file (xls extension but is actually a CSV) to DIRECTORY_DATA folder
  Edit grade centre spreadsheet (save as CSV with same xls extension):
    - if you have extra calculations to do in the sheet; probably best to copy from a master Excel file (paste values)
    - leave ONLY desired assessments, in desired order, with desired names
    - replace the id after pipe | with assessment weighting, e.g., 20% = "Assignment 1 [Total Pts: 100 Score] |20"
Annoying text like "Ready to Post(" or anything with parentheses will be stripped out automatically

* Put all related student class list CSV files in DIRECTORY_DATA
  That is, if your LearnJCU site is a merge of 3 subjects, you can include all 3 subjects to be processed at once :)
  The program will process each of the CSV files using the one LearnJCU Grade Centre file
  so do not have any other CSV files in this folder
* Run
    Output files will go into DIRECTORY_OUTPUT_NAME folder,
    ***Names will match class list CSV file names + "Results" - csv & xlsx
* Edit these to complete and confirm all is OK
  If you need to update small things, just do so and re-run to produce new output files
  Add your name to results sheet, other details should be automatic
  If you choose not to output a CSV, you can just export one manually (save as)
  from the xlsx file - useful if you want to update the results spreadsheet

TODO: consider rewriting with numpy/pandas
https://stackoverflow.com/questions/16503560/read-specific-columns-from-a-csv-file-with-csv-module
# TODO update current assumption for campus
# Assume filename format is like "CSV Report - Cairns_SpkCP1401Ver3-Results.xlsx"
"""
import csv
import os
import warnings
import openpyxl
import openpyxl.utils
import xlwings

# User-defined constants for preferences
WILL_CREATE_CSV = False  # Set this to True/False if you want/don't grade CSVs made
DIRECTORY_DATA = 'data/subject_results'
DIRECTORY_OUTPUT_NAME = 'output'
FILE_RESULTS_BLANK = 'Blank-Results.xlsx'
FILE_GRADE_CENTRE = 'learnjcu_grade_centre.xls'

# Internal constants that might change if spreadsheet structure changes
# Some row and column values are magic numbers; could be extracted as constants in future
SHEET_CLASS = 'StudentOne'
SHEET_RESULTS = 'RawResults'
# If Overall Grade is set, this will be the start text of the last column, otherwise the
LAST_HEADING_BEFORE_ASSESSMENTS_OVERALL = "Overall Grade"
LAST_HEADING_BEFORE_ASSESSMENTS_NO_OVERALL = 'Child Subject ID'  # for grade centre exports from merged LearnJCU sites
LAST_HEADING_BEFORE_ASSESSMENTS_NO_MERGE = 'Availability'  # for grade centre exports from non-merged LearnJCU sites
COLUMN_GRADE_CENTRE_ID = 3  # 3  # Numbered from 0 (csv/list) BUT!! Not always the same, so using dynamic (below)
STUDENT_ID_HEADING_TEXT = "Student ID"
ROW_CLASS_FIRST_STUDENT = 3
COLUMN_CLASS_SUBJECT_CODE = 6
COLUMN_CLASS_SUBJECT_NAME = 9
COLUMN_CLASS_YEAR = 10
COLUMN_CLASS_STUDY_PERIOD = 11
COLUMN_CLASS_STUDENT_ID = 13  # N
COLUMN_RESULTS_FIRST_ASSESSMENT = 5  # Numbered from 1 (openpyxl)
COLUMN_RESULTS_FINAL_GRADE_LETTER = 'M'  # O for U/S/X grades as in 2020-1, M for normal
ROW_RESULTS_FIRST_STUDENT = 13


def main():
    """Create subject results comprehensive spreadsheet."""
    warnings.filterwarnings("ignore")
    os.chdir(DIRECTORY_DATA)

    # create output folder if needed
    try:
        os.mkdir(f"{DIRECTORY_OUTPUT_NAME}")
    except FileExistsError:
        pass

    # get all relevant data from LearnJCU Grade Centre file (may be for multiple offerings)
    try:
        assessments, student_grade_centre_rows = get_assessments(FILE_GRADE_CENTRE)
    except RuntimeError as error:
        print(error)
        return
    print(f"Got {len(assessments)} assessments from {FILE_GRADE_CENTRE}")

    # get input filenames (each CSV is a different subject offering)
    class_list_filenames = [filename for filename in os.listdir('.') if filename.lower().endswith(".csv")]

    for class_list_filename in class_list_filenames:
        # Create filenames
        # TODO: Is there a way to standardise, or to determine the format from the existing name?

        # Ver 1 - Example filename: 'CP3402 TSV SP1 2023.Csv'
        # 2023-3- Example filename: 'CP1404 2023 CNS TR3.Csv'
        relevant_parts = class_list_filename.split()
        subject_code = relevant_parts[0]
        campus = relevant_parts[1]  # This changes sometimes :(

        # Ver 2 - Example filename: 'CSV Report - Cairns_SpkCP1401Ver3-Results.xlsx'
        # relevant_parts = class_list_filename.split()[-1].split('_')
        # campus = relevant_parts[0]
        # subject_code = relevant_parts[1][3:9]

        output_filename_base = f"{subject_code} {campus}"

        students_class_list = get_students(class_list_filename)
        print(f"Got {len(students_class_list)} students from {class_list_filename}")

        # assessments = [(7, 'Assignment 1', 100.0, 10), (8, 'Assignment 2', 100.0, 20)]
        student_results = get_student_results(student_grade_centre_rows, students_class_list, assessments)

        if len(students_class_list) != len(student_results):
            print("ERROR: Student results don't match class list")
            print("Trying with current data")
        print(f"Got {len(student_results)} students' results from {FILE_GRADE_CENTRE}")

        write_results(student_results, students_class_list, assessments, output_filename_base)
        print(f"Results spreadsheet created: {DIRECTORY_DATA}/{DIRECTORY_OUTPUT_NAME}/{output_filename_base}.xlsx")
        if WILL_CREATE_CSV:
            write_csv(output_filename_base)
            print(f"Grades CSV spreadsheet created: {DIRECTORY_DATA}/{DIRECTORY_OUTPUT_NAME}/{output_filename_base}.csv")


def get_students(input_filename):
    """Get students as a list of tuples from the class list."""
    students = []  # list of rows/lists
    input_file = open(input_filename, 'r')
    input_file.readline()  # Ignore first header line
    input_file.readline()  # Ignore second header line
    reader = csv.reader(input_file)
    for row in reader:
        student = row
        students.append(student)
    input_file.close()
    return students


def get_assessments(input_filename):
    """Extract assessment details from LearnJCU Grade Centre sheet."""
    # File from LearnJCU is UTF-16 encoded tab-delimited CSV with XLS extension
    try:
        input_file = open(input_filename, 'r', encoding='utf-16')
        rows = list(csv.reader(input_file, delimiter="\t"))
        input_file.close()
    except UnicodeError:
        raise RuntimeError(f"ERROR: Incorrect format; create {FILE_GRADE_CENTRE} again from download without changing format (copy and paste values into downloaded file)")
    headings = rows[0]

    # Get first assessment index based on possibilities of Grade Centre exports
    last_indexes = [i for i, heading in enumerate(headings) if heading.startswith(LAST_HEADING_BEFORE_ASSESSMENTS_OVERALL)]
    if last_indexes:
        first_assessment_index = last_indexes[0] + 1
    else:
        # Assessments start after the column with heading LAST_HEADER_BEFORE_ASSESSMENTS
        try:
            first_assessment_index = headings.index(LAST_HEADING_BEFORE_ASSESSMENTS_NO_OVERALL) + 1
        except ValueError:
            first_assessment_index = headings.index(LAST_HEADING_BEFORE_ASSESSMENTS_NO_MERGE) + 1
    assessment_headings = headings[first_assessment_index:]
    assessments = []
    # Assessment data looks like (last value needs to be set manually in spreadsheet):
    # heading = "Assignment 1 - Movies to Watch 1.0 [Total Pts: 100 Score] |20"
    # heading = "Pracs [Total Pts: up to 30 Score] |10"
    for i, heading in enumerate(assessment_headings):
        try:
            heading = heading.replace("up to ", "")  # for total columns
            parts = heading.split(" [Total Pts: ")
            title = parts[0]
            sub_parts = parts[1].split()
            score = float(sub_parts[0])
            weight = int(sub_parts[-1].strip('|'))
            assessments.append((first_assessment_index + i, title, score, weight))
        except IndexError:
            raise RuntimeError(f"ERROR with column data/structure in {DIRECTORY_DATA}/{input_filename}")
    # Return both processed assessments and all rows (including headings)
    return assessments, rows


def get_student_results(student_grade_centre_rows, students, assessments):
    """Get relevant student results from raw data in form based on assessments."""
    student_results = []
    # Get column for Student ID from grade centre (it's not always the same!)
    column_grade_centre_id = student_grade_centre_rows[0].index(STUDENT_ID_HEADING_TEXT)
    # Now that we have the relevant heading detail, remove heading row
    student_grade_centre_rows = student_grade_centre_rows[1:]

    grade_centre_ids = [row[column_grade_centre_id] for row in student_grade_centre_rows]
    for student in students:
        student_id = student[COLUMN_CLASS_STUDENT_ID]
        student_name = student[COLUMN_CLASS_STUDENT_ID + 1]
        student_result = [student_id, student_name]
        try:
            index = grade_centre_ids.index(student_id)
        except ValueError:
            print(f"ERROR?: Student: {student_name} ({student_id}) not in the LearnJCU Grade Centre sheet")
            # Need to store blank results so writing to spreadsheet works correctly
            student_result += [None] * len(assessments)
            student_results.append(student_result)
            continue
        for assessment in assessments:
            cell_value = student_grade_centre_rows[index][assessment[0]]
            try:
                score = float(cell_value)
            except ValueError:
                # Some cells have values like "In Progress(59.50)", which should just be the number
                if "(" in cell_value:
                    cell_value = cell_value[cell_value.find("(") + 1:-1]
                    score = float(cell_value)
                else:
                    # Replace non-scores like 'In Progress' or 'Needs Grading' with blanks
                    score = None
            student_result.append(score)
        student_results.append(student_result)
    return student_results


def write_results(student_results, class_list, assessments, output_filename_base):
    """Write student details and scores to results spreadsheet."""

    # Add student scores to results data sheet
    workbook = openpyxl.load_workbook(filename=FILE_RESULTS_BLANK)
    sheet = workbook[SHEET_CLASS]
    # Write all data from class list to results StudentOne sheet
    for i, student in enumerate(class_list):
        current_row = ROW_CLASS_FIRST_STUDENT + i
        for j, value in enumerate(student, 1):
            sheet.cell(row=current_row, column=j, value=value)
        # 2023 version: don't write formula; it's already there
        # # Write reference to grade using INDEX-MATCH so sorting results doesn't break formulas
        # reference_row = ROW_CLASS_FIRST_STUDENT + i
        # column_number_grade = openpyxl.utils.cell.column_index_from_string(COLUMN_RESULTS_FINAL_GRADE_LETTER)
        # # + 1 since csv/pyopenxl have different starting indexes
        # column_letter_id = openpyxl.utils.cell.get_column_letter(COLUMN_CLASS_STUDENT_ID + 1)
        # formula = f"=INDEX({SHEET_RESULTS}!A${ROW_RESULTS_FIRST_STUDENT}:Q$320,MATCH({column_letter_id}{reference_row},{SHEET_RESULTS}!B${ROW_RESULTS_FIRST_STUDENT}:B$320,0),{column_number_grade})"
        # sheet.cell(row=current_row, column=COLUMN_CLASS_STUDENT_ID + 3, value=formula)

    # Add formulas to raw results sheet to refer to student ID and name
    sheet = workbook[SHEET_RESULTS]
    for i in range(len(class_list)):
        current_row = ROW_RESULTS_FIRST_STUDENT + i
        reference_row = ROW_CLASS_FIRST_STUDENT + i
        sheet.cell(row=current_row, column=2, value=f"={SHEET_CLASS}!N{reference_row}")
        sheet.cell(row=current_row, column=3, value=f"={SHEET_CLASS}!O{reference_row}")

    # Get and write subject information to Results
    # See filename format in top CONSTANTS
    subject_name = class_list[0][COLUMN_CLASS_SUBJECT_NAME]
    campus = output_filename_base.split()[1]  # Based on assumptions (see comments at top)
    subject_code = class_list[0][COLUMN_CLASS_SUBJECT_CODE]
    year = class_list[0][COLUMN_CLASS_YEAR]
    study_period = class_list[0][COLUMN_CLASS_STUDY_PERIOD]
    sheet.cell(row=1, column=3, value=subject_name)
    sheet.cell(row=2, column=3, value=campus)
    # row 3 is "Mode" - not in the CSV
    sheet.cell(row=4, column=3, value=year)
    sheet.cell(row=5, column=3, value=study_period)
    sheet.cell(row=6, column=3, value=subject_code)  # but there's a formula already in the results sheet for this
    # Update filename now that we know details
    output_filename_base = f"{output_filename_base} {year} {study_period}-Results"

    # Add assessment headings
    # Assume spreadsheet structure is based on 13=ROW_FIRST_RAW_DATA, 10=%, 11=Out Of, 12=Title
    # assessment looks like (7, 'Assignment 1', 100.0, 20)
    row_weight = ROW_RESULTS_FIRST_STUDENT - 3
    row_out_of = ROW_RESULTS_FIRST_STUDENT - 2
    row_title = ROW_RESULTS_FIRST_STUDENT - 1
    for i, assessment in enumerate(assessments):
        # Note: staff need to enter weight in Grade Centre file first
        weight_to_write = assessment[3] / 100  # Need 0-1 instead of 0-100 for % in spreadsheet
        sheet.cell(row=row_weight, column=COLUMN_RESULTS_FIRST_ASSESSMENT + i, value=weight_to_write)
        sheet.cell(row=row_out_of, column=COLUMN_RESULTS_FIRST_ASSESSMENT + i, value=assessment[2])
        sheet.cell(row=row_title, column=COLUMN_RESULTS_FIRST_ASSESSMENT + i, value=assessment[1])

    # Add assessment scores to RawResults sheet
    for i, current_student_results in enumerate(student_results):
        current_row = ROW_RESULTS_FIRST_STUDENT + i
        # Write just row/student reference number (1, 2...)
        sheet.cell(row=current_row, column=1, value=i + 1)

        for j, score in enumerate(current_student_results[2:]):
            if score is None:  # Don't write blanks
                continue
            current_column = COLUMN_RESULTS_FIRST_ASSESSMENT + j
            sheet.cell(row=current_row, column=current_column, value=score)

    # Write blank row so that sorting works in Excel
    for column_number in range(1, 20):  # About 16 columns - clear them all
        sheet.cell(row=current_row + 1, column=column_number, value="")

    workbook.save(filename=f"{DIRECTORY_OUTPUT_NAME}/{output_filename_base}.xlsx")


def write_csv(filename_base):
    """Write just CSV file needed for results submission."""
    # TODO: fix this as filename has changed (2023)
    input_filename = f"{filename_base}.xlsx"
    output_filename = f"{filename_base}.csv"
    # openpyxl does not evaluate formulas, so use xlwings
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(f"{DIRECTORY_OUTPUT_NAME}/{input_filename}")
    sheet = excel_book.sheets[SHEET_CLASS]
    with open(f"{DIRECTORY_OUTPUT_NAME}/{output_filename}", 'w', newline="") as output_file:
        writer = csv.writer(output_file)
        for row in sheet.range('A1').current_region.value:
            writer.writerow(row)
    excel_book.save()
    excel_book.close()
    excel_app.quit()


main()
