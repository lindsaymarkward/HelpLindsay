"""Get CPXXXX assessment items from StudyFinder websites.

Manually set column C "Check"
=IF(AND(ISBLANK(H2), ISBLANK(M2), ISBLANK(R2), ISBLANK(W2)), 0, 1)
"""

from openpyxl import Workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright

YEAR = 2024
OUTPUT_EXCEL_FILENAME = "output/subject_assessment.xlsx"


def main():
    """Load CP assessment details from the web and save them to an Excel file."""
    subject_to_items = {}
    year_to_get = YEAR
    # subjects = ['CP1401', 'CP1402', 'CP5638']  # Short list just for testing
    subjects = load_subjects()
    print("Processing assessment items for", subjects)

    for subject in subjects:
        subject_name, assessment_text = extract_assessment(year_to_get, subject)
        items = extract_items(assessment_text)
        subject_to_items[subject] = [subject_name, items]
    # print(subject_to_items)
    write_spreadsheet(subject_to_items)


def extract_assessment(year, subject_code):
    """Download and extract one subject's assessment details from JCU CSDB website."""
    url = f"https://apps.jcu.edu.au/subjectsearch/#/subject/{year}/{subject_code}"
    print("Getting", subject_code)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(url)
        # Wait for the page to load completely
        page.wait_for_selector("h3:has-text('Subject Assessment')")

        # Get the subject name
        title = page.query_selector("h2").text_content()
        subject_name = title.split(" - ")[1]
        # Locate the <h3> element and then find the following <ul>
        assessment_content = page.query_selector("h3:has-text('Subject Assessment') + ul")
        if assessment_content:
            assessment_text = assessment_content.inner_text()
        else:
            assessment_text = "No assessment information found."
        browser.close()
    return subject_name, assessment_text


def load_subjects():
    """Load all subject codes from text file."""
    file_in = open("data/all_subjects.txt")
    subjects = [line.strip() for line in file_in]
    file_in.close()
    return subjects


def extract_items(block):
    """Extract assessment items from text block as list of tuples."""
    items = []
    for line in block.split('\n'):
        parts = line.split(' - ')
        try:
            mode = parts[0]
            weight = int(parts[1].strip('(').strip('%)'))
            group = parts[2]
            items.append((mode, weight, group))
        except ValueError:
            print(f"Fixing invalid int: {line}")
            # This is almost surely a "-" in the assessment item name (only one for CP5638 at time of testing)
            mode = f"{parts[0]} - {parts[1]}"
            del parts[1]
            weight = int(parts[1].strip('(').strip('%)'))
            group = parts[2]
            items.append((mode, weight, group))
        except IndexError:
            print(f"ERROR with: {line}")
    return items


def write_spreadsheet(subject_to_items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Assessment"

    # Write header titles, named to support Excel data table where column headers must be unique
    headers = ["Code", "Subject Title", "Check",
               "Item1", "Mode1", "Weight1", "Group1", "Moderation1",
               "Item2", "Mode2", "Weight2", "Group2", "Moderation2",
               "Item3", "Mode3", "Weight3", "Group3", "Moderation3",
               "Item4", "Mode4", "Weight4", "Group4", "Moderation4"]

    # Write headers to the first row
    for col_number, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_number, value=header)
        cell.font = Font(bold=True)

    # Write per-subject assessment items
    row = 2
    for subject_code, items in subject_to_items.items():
        subject_name, assessment_items = items
        col = 1
        ws.cell(row=row, column=col, value=subject_code)
        col += 1
        ws.cell(row=row, column=col, value=subject_name)
        col += 3
        for item in assessment_items:
            for element in item:
                ws.cell(row=row, column=col, value=element)
                col += 1
            col += 2  # Leave room for comment and a gap between assessment items
        row += 1
    wb.save(OUTPUT_EXCEL_FILENAME)


def write_spreadsheet_2(all_subject_details):
    # each value contains (a list of items, prerequisite string)
    workbook = openpyxl.load_workbook(filename=OUTPUT_EXCEL_FILENAME)
    sheet = workbook['Assessment-Mapping']
    row = 12  # first row for assessment items
    column = 2
    for subject, items in all_subject_details.items():
        # sheet.cell(row=row, column=column, value=subject)
        prerequisite = items[1]
        items = items[0]  # effectively rename as (assessment) items, without prerequisite
        item_row = row
        for item_number, item in enumerate(items):
            name, weight = item
            try:
                weight = int(weight)
            except ValueError:
                pass
            sheet.cell(row=item_row + item_number, column=column, value=name)
            sheet.cell(row=item_row + item_number, column=column + 3, value=weight)
        sheet.cell(row=24, column=column, value=prerequisite)

        column += 4  # distance to next subject (4 pieces of data per assessment)
    # row += 7  # move down to write next year's items
    # column = 2
    workbook.save(filename=f"output/temp.xlsx")


def run_tests():
    block = """Performance/Practice/Product > Software development/creation - (60%) - Individual
Performance/Practice/Product > Practical assessment/practical skills demonstration - (40%) - Individual"""
    items = extract_items(block)
    print(items)


# run_tests()
main()
