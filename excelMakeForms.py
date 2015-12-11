import openpyxl

__author__ = 'sci-lmw1'

groupsWB = openpyxl.load_workbook('CP1406Groups.xlsx')
groupsSheet = groupsWB.get_sheet_by_name("Cairns")
# campuses = groupsWB.get_sheet_names()
campus = groupsSheet.title

# print(groupsSheet.cell(row=2, column=1).value)
# print(groupsSheet.max_row)

data = {}

# open template sheet - rewrite over, save each time
toWriteWB = openpyxl.load_workbook('CP1406PeerAssessmentForm.xlsx')
sheet = toWriteWB.active

for i in range(1, groupsSheet.max_row):
    groupCell, usernameCell, lastNameCell, firstNameCell = groupsSheet.rows[i]
    # print(group.value, firstName.value, lastName.value)

    # add data to dictionary
    if groupCell.value not in data:
        data[groupCell.value] = [(firstNameCell.value, lastNameCell.value)]
    else:
        data[groupCell.value].append((firstNameCell.value, lastNameCell.value))

# write to sheet
# print("?? ", data)
for group, names in data.items():
    sheet.title = 'Group ' + str(group)
    sheet['B2'] = group
    for personNumber, nameParts in enumerate(names):
        sheet.cell(row=9 + personNumber, column=1).value = nameParts[0]
        sheet.cell(row=9 + personNumber, column=2).value = nameParts[1]
    # print("!!", group, names)
    for personNumber, nameParts in enumerate(names):
        toWriteWB.save(campus + str(group) + nameParts[0] + nameParts[1] + 'PA.xlsx')


# print(data)

# newWB = openpyxl.Workbook()
# newWB.title = 'Group n'
# newSheet = newWB.active
# newSheet['A1'] = "O YAY!"
# newWB.save("FTest.xlsx")



